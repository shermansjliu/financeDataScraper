from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from selenium import webdriver
from typing import Union

from selenium.webdriver.support.ui import WebDriverWait
from Info import Info


class FinanceBot:
    def __init__(self, workbook_name, start_row, start_col):
        self.driver = webdriver.Chrome("C:\chromedriver")
        self.info = {info.value: 0. for info in Info}
        self.wait = WebDriverWait(self.driver, 10)
        self.wb = load_workbook(filename=f"{workbook_name}.xlsx")
        self.wb_name = workbook_name
        self.start_col = start_col
        self.start_row = start_row
        # TODO initialize new worksheet
        self.font = Font(b=False, size=11)

    def initialize_worksheet(self):
        sheet = self.wb.active
        column_names = list(self.info.keys())
        column_names.sort()

        # Clear instruction columns
        sheet.delete_cols(self.starting_column, len(self.info) + self.starting_column)
        # First two rows are company ids
        col_i = self.starting_column

        for name in column_names:
            header_cell = sheet.cell(row=1, column=col_i)
            header_cell.value = name
            header_cell.font = self.font
            if name == Info.COMPANY_NAME.value:
                sheet.column_dimensions[header_cell.column_letter].width = 40
            elif name == Info.EBIT.value:
                sheet.column_dimensions[header_cell.column_letter].width = 30
            else:
                sheet.column_dimensions[header_cell.column_letter].width = len(name) * 1.2
            col_i += 1

        # Take categories and iterate over them.
        self.wb.save(filename=f"{self.wb_name}.xlsx")

        # PATH = "/Users/work/Documents/projects/chromedriver"

    def extract_income_statement_page(self, company_code: int) -> None:
        """
        Extracts pertinent data from the income statement page
        """
        self.driver.get(
            f"https://www.wsj.com/market-data/quotes/HK/XHKG/{company_code}/financials/annual/income-statement"
        )
        sales_revenue = self.get_latest_cell_value(Info.S_R.value)
        ebit = self.get_latest_cell_value(Info.EBIT.value)
        self.info[Info.S_R.value] = sales_revenue
        self.info[Info.EBIT.value] = ebit

    def get_latest_cell_value(self, cell_name: str) -> Union[float, str]:

        elmts = self.driver.find_elements_by_xpath(f"//td[text()='{cell_name}']/following-sibling::td[1]")
        # If doesn't exist or emtpy string return N/A
        if not elmts:
            # print(f"{cell_name}: N/A")
            return "N/A"

        elmt = elmts[0]

        if elmt.text == "":
            # print(f"{cell_name}: N/A")
            return "N/A"

        # If text == dash return 0
        elif elmt.text == "-":
            # print(f"{cell_name}: 0")
            return 0.

        # Remove comma b/c Python cannot convert float of string numbers with commas e.g ("100,000")
        elmt_text = elmt.text.replace("(", "").replace(")", "").replace(",", "")
        # print(f"{cell_name}: {elmt_text}")
        return float(elmt_text)

    def extract_balance_sheet_page(self, company_code: int) -> None:

        self.driver.get(
            f"https://www.wsj.com/market-data/quotes/HK/XHKG/{company_code}/financials/annual/balance-sheet")

        self.info[Info.CASH_ST_INVESTMENTS.value] = self.get_latest_cell_value(Info.CASH_ST_INVESTMENTS.value)
        self.info[Info.TCA.value] = self.get_latest_cell_value(Info.TCA.value)

        # Expand Liabilities & Shareholders' Equity Section
        liabilities_btn = self.driver.find_elements_by_xpath("//h2[contains(text(),Liabilities)]/parent::div")
        if liabilities_btn:
            liabilities_btn[1].click()

        self.info[Info.TOTDebt.value] = self.get_latest_cell_value(Info.TOTDebt.value)
        self.info[Info.LTDebt.value] = self.get_latest_cell_value(Info.LTDebt.value)
        self.info[Info.CPLTDebt.value] = self.get_latest_cell_value(Info.CPLTDebt.value)
        self.info[Info.STDebt.value] = self.get_latest_cell_value(Info.STDebt.value)
        self.info[Info.NETPPEQ.value] = self.get_latest_cell_value(Info.NETPPEQ.value)
        self.info[Info.TCL.value] = self.get_latest_cell_value(Info.TCL.value)

    def page_not_found(self)->bool:
        pg_not_found_elmt = self.driver.find_elements_by_xpath("//span[text()='Page Not Found']")
        if len(pg_not_found_elmt) > 0:
            return True
        else:
            return False

    def extract_finance_data_from_company(self, company_code: int) -> None:

        self.driver.get(
            f"https://www.wsj.com/market-data/quotes/HK/XHKG/{company_code}/financials"
        )

        if self.page_not_found():
            for key in self.info:
                self.info[key] = "N/A"
            return

        company_name = self.driver.find_element_by_class_name("companyName").text
        company_name = company_name[
                       : len(company_name) - 2
                       ]  # Remove Full stop and white space from title
        self.info[Info.COMPANY_NAME.value] = company_name
        # Extract data from financials page
        entrprse_value_to_sales_elmts = self.driver.find_elements_by_xpath(
            "//span[text()='Enterprise Value to Sales']/following-sibling::span[1]/span")
        if not entrprse_value_to_sales_elmts == "" or entrprse_value_to_sales_elmts[0] == "":
            self.info[Info.EVTS.value] = "N/A"
        elif entrprse_value_to_sales_elmts[0] == "-":
            self.info[Info.EVTS.value] = 0.
        else:
            self.info[Info.EVTS.value] = float(entrprse_value_to_sales_elmts[0].text)

        # Visit Income Statement
        self.extract_income_statement_page(company_code)

        # Visit Balance Sheet Page
        self.extract_balance_sheet_page(company_code)

    def extract_finance_data(self) -> None:
        sheet = self.wb.active

        company_row = self.start_row

        while sheet.cell(row=company_row, column=1).value is not None:
            company_id = int(sheet.cell(row=company_row, column=1).value)
            self.extract_finance_data_from_company(company_id)
            for col in range(self.start_col, self.start_col + len(self.info)):
                col_name = sheet.cell(row=1, column=col).value
                cell = sheet.cell(row=company_row, column=col)
                # print(f"{cell.column_letter}:{company_row}, {col_name}: {self.info.get(col_name)}")
                cell.font = self.font
                cell.alignment = Alignment(horizontal="left")
                # print(f"f{col_name}, ")
                cell.value = self.info[col_name]
            company_row += 1
            self.wb.save(filename=f"{self.wb_name}.xlsx")
        # self.wb.save(filename=f"{self.wb_name}.xlsx")
        self.driver.quit()
