from openpyxl import Workbook, load_workbook
from selenium import webdriver


column_names=[
"Enterprise Value to Sales",
"EBIT",
"Sales/Revenue",
"cash",
"total current assets",
"short term debt",
"current portion of long term debt",
"total current liabilities",
"Net Property, Plant & Equipment"]

def format_worksheet():
  wb = load_workbook(filename="copy_hk_stock.xlsx")
  # print(wb.sheetnames)
  sheet = wb.active
  sheet["B1"] = "Hello"


  # Clear columns
  sheet.delete_cols(3,4)
  #Populate column names
  col_i = 3
  for name in column_names:
    sheet.cell(row=1, column=col_i).value = name
    col_i += 1

  #Take categories and iterate over them. 
  wb.save(filename="copy_hk_stock.xlsx")


financials_url = 'https://www.wsj.com/market-data/quotes/HK/XHKG/1810/financials'
PATH = "/Users/work/Documents/projects/chromedriver"

driver = webdriver.Chrome(PATH)
driver.get(financials_url)

def extract_finance_data(company_code):
  driver.get(f"https://www.wsj.com/market-data/quotes/HK/XHKG/{code}/financials")

 
  
  company_name = driver.find_element_by_class_name('companyName')
  etrprse_val_to_sales = find_element_by_link_text('Enterprise Value to Sales')
  print(company_name)
  print(etrprise_val_to_sales)


extract_finance_data(1810)



if __name__ == "__main__":
  # format_worksheet()
  extract_finance_data()

